from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
import asyncio
import pytest

from gst_automation.gst.download_verifier import Gstr2bDownloadVerifier


def _write_xlsx(path: Path, rows: int = 5) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["a", "b", "c"])
    for i in range(rows):
        ws.append([i, i + 1, i + 2])
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def test_download_verifier_accepts_valid_xlsx(tmp_path: Path) -> None:
    p = tmp_path / "GSTR2B_24ABCDE1234F1Z5_2026-05.xlsx"
    _write_xlsx(p)
    res = Gstr2bDownloadVerifier().verify(p)
    assert res.ok
    assert res.sha256_hex
    assert res.byte_size > 0
    assert res.sheet_names


def test_download_verifier_flags_truncated_xlsx(tmp_path: Path) -> None:
    p = tmp_path / "bad.xlsx"
    _write_xlsx(p)
    data = p.read_bytes()
    p.write_bytes(data[:100])  # truncate
    res = Gstr2bDownloadVerifier().verify(p)
    assert not res.ok
    assert res.classification in {"retryable", "corrupt"}


def test_client_master_parser_counts_rows_with_string_client_id(tmp_path: Path) -> None:
    from gst_automation.clients.excel_parser import COLUMNS, ClientMasterParser

    p = tmp_path / "client_master.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "CLIENT_MASTER"
    ws.append([h for h in COLUMNS])
    ws.append(
        [
            "CLIENT001",
            "Jay Ambe Products",
            "24AAEFJ7740C1ZJ",
            "JAPAAEFJ7740C",
            "password",
            "test@gmail.com",
            "2025-26",
            True,
            "MEDIUM",
            "JAG",
            16,
            "notes",
        ]
    )
    wb.save(p)

    parsed = ClientMasterParser(path=p).parse()
    assert parsed.ok
    assert len(parsed.rows) == 1


def test_client_master_import_preview_returns_rows(tmp_path: Path) -> None:
    from gst_automation.clients.excel_parser import COLUMNS
    from gst_automation.clients.import_pipeline import ClientImportPipeline
    from gst_automation.core.settings import Settings

    p = tmp_path / "client_master.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "CLIENT_MASTER"
    ws.append([h.upper() for h in COLUMNS])  # header normalization should accept uppercase
    ws.append(
        [
            "CLIENT001",
            "Jay Ambe Products",
            "24AAEFJ7740C1ZJ",
            "JAPAAEFJ7740C",
            "password",
            "test@gmail.com",
            "2025-26",
            "TRUE",
            "MEDIUM",
            "JAG",
            "16",
            "notes",
        ]
    )
    wb.save(p)

    # Settings requires DATABASE_URL; set a dummy for this unit test.
    monkeypatch = pytest.MonkeyPatch()
    monkeypatch.setenv("DATABASE_URL", "postgresql+asyncpg://x:y@localhost:5432/z")
    Settings.load.cache_clear()  # type: ignore[attr-defined]
    settings = Settings.load()
    rep = asyncio.run(ClientImportPipeline(settings=settings).import_xlsx(None, path=p, dry_run=True))
    monkeypatch.undo()
    assert rep.ok
    assert int(rep.summary["total_rows"]) == 1
    assert len(rep.preview) == 1

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


@dataclass(frozen=True, slots=True)
class ClientMasterTemplate:
    """Generates `client_master.xlsx` for production client onboarding."""

    path: Path

    def write(self) -> None:
        wb = Workbook()

        headers: list[str] = [
            "client_id",
            "client_name",
            "gstin",
            "username",
            "password",
            "client_email",
            "financial_year",
            "active",
            "priority",
            "tags",
            "preferred_run_window",
            "notes",
        ]

        required_cols = {
            "client_name",
            "gstin",
            "username",
            "password",
            "client_email",
            "financial_year",
            "active",
            "priority",
            "preferred_run_window",
        }

        header_fill_required = PatternFill("solid", fgColor="FDE68A")  # amber-200
        header_fill_optional = PatternFill("solid", fgColor="E5E7EB")  # gray-200
        header_font = Font(color="111827", bold=True)  # gray-900
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Sheet 1: CLIENT_MASTER
        ws = wb.active
        ws.title = "CLIENT_MASTER"
        ws.append(headers)

        for idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=idx, value=h)
            cell.fill = header_fill_required if h in required_cols else header_fill_optional
            cell.font = header_font
            cell.alignment = header_alignment
            ws.column_dimensions[get_column_letter(idx)].width = max(14, min(34, len(h) + 2))

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
        ws.row_dimensions[1].height = 24

        col = {h: i + 1 for i, h in enumerate(headers)}
        max_rows = 1000

        # Validations.
        dv_active = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=False)
        dv_priority = DataValidation(type="list", formula1='"HIGH,MEDIUM,LOW"', allow_blank=False)
        dv_window = DataValidation(type="list", formula1='"15,16,17,18,19,20"', allow_blank=False)
        ws.add_data_validation(dv_active)
        ws.add_data_validation(dv_priority)
        ws.add_data_validation(dv_window)

        dv_active.add(f"{get_column_letter(col['active'])}2:{get_column_letter(col['active'])}{max_rows}")
        dv_priority.add(f"{get_column_letter(col['priority'])}2:{get_column_letter(col['priority'])}{max_rows}")
        dv_window.add(
            f"{get_column_letter(col['preferred_run_window'])}2:{get_column_letter(col['preferred_run_window'])}{max_rows}"
        )

        # Comments / tooltips (never put secrets in comments).
        ws.cell(row=1, column=col["client_id"]).comment = Comment(
            "Optional for new clients. Leave blank to auto-generate a stable id during import. "
            "Keep the generated id unchanged for future updates.",
            "GST Automation",
        )
        ws.cell(row=1, column=col["gstin"]).comment = Comment(
            "GSTIN rules: 15 chars, uppercase, alphanumeric, valid GSTIN pattern.\nExample: 24ABCDE1234F1Z5",
            "GST Automation",
        )
        ws.cell(row=1, column=col["client_email"]).comment = Comment(
            "Comma-separated emails supported (e.g. a@x.com, b@y.com). Each email must be valid.",
            "GST Automation",
        )
        ws.cell(row=1, column=col["password"]).comment = Comment(
            "Security: Passwords are stored ONLY in the Vault. "
            "They are never stored in plaintext DB rows. Do not share externally.",
            "GST Automation",
        )
        ws.cell(row=1, column=col["preferred_run_window"]).comment = Comment(
            "Preferred execution start hour (24h) for overnight batches. Supported: 15–20.",
            "GST Automation",
        )
        ws.cell(row=1, column=col["priority"]).comment = Comment(
            "Scheduling priority: HIGH runs earlier, MEDIUM normal, LOW last.",
            "GST Automation",
        )

        for r in range(2, max_rows + 1):
            ws.cell(row=r, column=col["active"]).alignment = Alignment(horizontal="center")
            ws.cell(row=r, column=col["priority"]).alignment = Alignment(horizontal="center")
            ws.cell(row=r, column=col["preferred_run_window"]).alignment = Alignment(horizontal="center")
            ws.cell(row=r, column=col["financial_year"]).alignment = Alignment(horizontal="center")

        # Sheet 2: INSTRUCTIONS (protected)
        ins = wb.create_sheet("INSTRUCTIONS")
        ins["A1"] = "GST Automation Platform — Client Master Workbook Instructions"
        ins["A1"].font = Font(bold=True, size=14)
        ins["A1"].alignment = Alignment(wrap_text=True)

        lines = [
            "Purpose: Use this workbook to onboard GST clients for automated GSTR-2B execution and overnight batch runs.",
            "",
            "How to add clients:",
            "1) Open the CLIENT_MASTER sheet.",
            "2) Add one row per GSTIN (one client = one GSTIN).",
            "3) Do not rename columns or change the header row.",
            "4) For new clients, you may leave client_id blank. Import will generate it and you should keep it for future updates.",
            "",
            "GSTIN formatting rules:",
            "- GSTIN must be 15 characters, uppercase, alphanumeric, and match standard GSTIN pattern.",
            "- Example GSTIN: 24ABCDE1234F1Z5",
            "",
            "Email formatting rules:",
            "- client_email can contain multiple emails separated by commas.",
            "- Each email must be valid (example: accounts@client.com, ops@client.com).",
            "",
            "Session reuse concept:",
            "- The system can reuse saved browser sessions to reduce logins.",
            "- If a session expires, an operator will refresh it once, and the next retries can proceed.",
            "",
            "Overnight batch execution concept:",
            "- Runs are typically scheduled in the afternoon/evening window (15–20) so outputs are ready for the next day.",
            "- preferred_run_window helps operations group clients into predictable time buckets.",
            "",
            "Execution window (15–20):",
            "- Allowed values: 15, 16, 17, 18, 19, 20 (24-hour clock).",
            "",
            "Retry window behavior:",
            "- If a run fails due to transient issues (timeouts, portal instability), the platform will retry.",
            "- If authentication is required, a supervised refresh is requested and the job becomes retryable.",
            "",
            "How passwords are secured:",
            "- Passwords are stored ONLY in the platform Vault.",
            "- The database stores only a secret reference (not the password itself).",
            "- Passwords must never be shared externally.",
            "",
            "Inactive clients:",
            "- active=FALSE means the client is onboarded but excluded from automated batches.",
            "- Inactive clients are not scheduled until set to TRUE.",
            "",
            "Meaning of priority:",
            "- HIGH: prioritize earlier execution; MEDIUM: default; LOW: deprioritized.",
            "",
            "How preferred_run_window is used:",
            "- Operations can schedule/monitor batches per hour bucket to balance load and handle auth refreshes.",
        ]
        for i, line in enumerate(lines, start=3):
            ins[f"A{i}"] = line
            ins[f"A{i}"].alignment = Alignment(wrap_text=True, vertical="top")
        ins.column_dimensions["A"].width = 140
        ins.protection.sheet = True
        ins.protection.enable()

        sample_rows = [
            [
                "",
                "Apex Components Pvt Ltd",
                "24ABCDE1234F1Z5",
                "apex.gst",
                "demo_password_1",
                "accounts@apex-components.example, ops@apex-components.example",
                "2025-26",
                "TRUE",
                "HIGH",
                "manufacturing, priority-client",
                16,
                "Runs weekly; prefers earlier window.",
            ],
            [
                "",
                "BlueSky Traders",
                "27PQRSX6789L1Z2",
                "bluesky.traders",
                "demo_password_2",
                "finance@bluesky-traders.example",
                "2025-26",
                "TRUE",
                "MEDIUM",
                "trading",
                18,
                "Comma-separated emails supported.",
            ],
            [
                "",
                "Cedar Services LLP",
                "29LMNOP4321Q1Z9",
                "cedar.services",
                "demo_password_3",
                "accounts@cedar-services.example",
                "2024-25",
                "FALSE",
                "LOW",
                "services, inactive",
                19,
                "Inactive client: onboarded but excluded from batches.",
            ],
        ]
        for r in sample_rows:
            ws.append(r)

        # Auto-adjust CLIENT_MASTER column widths based on content.
        for idx, h in enumerate(headers, start=1):
            letter = get_column_letter(idx)
            max_len = len(h)
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=idx, max_col=idx):
                val = row[0].value
                if val is None:
                    continue
                max_len = max(max_len, len(str(val)))
            ws.column_dimensions[letter].width = max(14, min(44, max_len + 2))

        self.path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(self.path)

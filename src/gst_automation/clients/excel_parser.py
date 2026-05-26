from __future__ import annotations

import re
import uuid
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from gst_automation.core.logging import get_logger


GSTIN_RE = re.compile(r"^[0-9]{2}[A-Z]{5}[0-9]{4}[A-Z]{1}[1-9A-Z]{1}Z[0-9A-Z]{1}$")
EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")

FINANCIAL_YEARS = {"2024-25", "2025-26", "2026-27"}
RUN_WINDOWS = {15, 16, 17, 18, 19, 20}
PRIORITIES = {"HIGH", "MEDIUM", "LOW"}

CLIENT_MASTER_SHEET = "CLIENT_MASTER"
LEGACY_CLIENTS_SHEET = "clients"

COLUMNS = [
    "client_id",
    "client_name",
    "gstin",
    "username",
    "password",
    "client_email",
    "financial_year",
    "active",
    "priority",
    "tags",
    "preferred_run_window",
    "notes",
]

REQUIRED_FIELDS = {
    "client_name",
    "gstin",
    "username",
    "password",
    "client_email",
    "financial_year",
    "active",
    "priority",
    "preferred_run_window",
}

logger = get_logger(__name__)


@dataclass(frozen=True, slots=True)
class RowError:
    row: int
    field: str
    message: str


@dataclass(frozen=True, slots=True)
class ParseResult:
    rows: list[dict[str, Any]]
    errors: list[RowError]

    @property
    def ok(self) -> bool:
        return not self.errors


@dataclass(frozen=True, slots=True)
class ClientRecord:
    client_name: str
    gstin: str
    username: str
    client_email: str
    financial_year: str
    active: bool
    priority: str
    tags: str | None
    preferred_run_window: int
    notes: str | None
    client_id: str | None = None
    password: str | None = None
    rownum: int | None = None


@dataclass(frozen=True, slots=True)
class TypedParseResult:
    records: list[ClientRecord]
    errors: list[RowError]
    headers: list[str]

    @property
    def ok(self) -> bool:
        return not self.errors


def _norm_str(v: object | None) -> str | None:
    if v is None:
        return None
    if isinstance(v, bool):
        return "TRUE" if v else "FALSE"
    s = str(v).strip()
    if s == "":
        return None
    return re.sub(r"\s+", " ", s)


def _norm_header(v: object | None) -> str | None:
    s = _norm_str(v)
    if s is None:
        return None
    # Normalize headers to match deterministic expected column keys.
    # Example: "Client Name" -> "client_name"
    s2 = re.sub(r"\s+", "_", s.strip().lower())
    return s2


def _parse_bool(v: str | None) -> bool | None:
    if v is None:
        return None
    u = v.strip().upper()
    if u in {"TRUE", "1", "YES", "Y"}:
        return True
    if u in {"FALSE", "0", "NO", "N"}:
        return False
    return None


def _split_emails(v: str) -> list[str]:
    return [p.strip() for p in v.split(",") if p.strip()]


@dataclass(frozen=True, slots=True)
class ClientMasterParser:
    path: Path

    def parse(self) -> ParseResult:
        wb = load_workbook(self.path)
        diags: list[RowError] = []
        try:
            diags.append(RowError(row=0, field="diagnostic", message=f"sheetnames={wb.sheetnames}"))
        except Exception:
            pass
        sheet = CLIENT_MASTER_SHEET if CLIENT_MASTER_SHEET in wb.sheetnames else None
        if sheet is None and LEGACY_CLIENTS_SHEET in wb.sheetnames:
            sheet = LEGACY_CLIENTS_SHEET
        if sheet is None:
            return ParseResult(
                rows=[],
                errors=[RowError(row=0, field="sheet", message=f"missing '{CLIENT_MASTER_SHEET}' sheet")],
            )

        ws = wb[sheet]
        headers = [c.value for c in ws[1]]
        headers_norm = [_norm_header(h) for h in headers]
        headers_norm = [h for h in headers_norm if h is not None]

        # For production onboarding, require exact column order for deterministic imports.
        if headers_norm != COLUMNS:
            return ParseResult(
                rows=[],
                errors=[
                    *diags,
                    RowError(
                        row=1,
                        field="header",
                        message=f"invalid header. Expected columns: {', '.join(COLUMNS)}",
                    )
                ],
            )

        errors: list[RowError] = []
        rows: list[dict[str, Any]] = []
        seen_gstin: set[str] = set()
        seen_username: set[str] = set()
        processed = 0
        skipped_empty = 0

        for i in range(2, ws.max_row + 1):
            raw = {h: ws.cell(row=i, column=idx + 1).value for idx, h in enumerate(headers_norm)}
            # Skip only fully-empty rows (handles None, empty strings, whitespace-only).
            if not any(_norm_str(v) is not None for v in raw.values()):
                skipped_empty += 1
                continue
            processed += 1

            def get(name: str) -> str | None:
                return _norm_str(raw.get(name))

            # Missing required fields.
            for f in REQUIRED_FIELDS:
                if get(f) is None:
                    errors.append(RowError(row=i, field=f, message="missing"))

            # Optional client_id; validate if present.
            client_id = get("client_id")
            # Accept non-UUID client_id (operator code) as valid; DB id is derived elsewhere.
            if client_id:
                try:
                    uuid.UUID(client_id)
                except Exception:
                    pass

            # GSTIN validation + duplicate detection.
            gstin = get("gstin")
            if gstin:
                gstin_u = gstin.upper().replace(" ", "")
                if len(gstin_u) != 15:
                    errors.append(RowError(row=i, field="gstin", message="must be 15 characters"))
                elif not GSTIN_RE.match(gstin_u):
                    errors.append(RowError(row=i, field="gstin", message="invalid GSTIN format"))
                if gstin_u in seen_gstin:
                    errors.append(RowError(row=i, field="gstin", message="duplicate GSTIN"))
                seen_gstin.add(gstin_u)

            # Username duplicate detection.
            username = get("username")
            if username:
                u = username.strip()
                if u in seen_username:
                    errors.append(RowError(row=i, field="username", message="duplicate username"))
                seen_username.add(u)

            # Email validation (comma-separated allowed).
            emails = get("client_email")
            if emails:
                parts = _split_emails(emails)
                if not parts:
                    errors.append(RowError(row=i, field="client_email", message="missing"))
                else:
                    for p in parts:
                        if not EMAIL_RE.match(p):
                            errors.append(RowError(row=i, field="client_email", message=f"invalid email: {p}"))

            fy = get("financial_year")
            if fy and fy not in FINANCIAL_YEARS:
                errors.append(RowError(row=i, field="financial_year", message="invalid financial_year"))

            active = get("active")
            if active and _parse_bool(active) is None:
                errors.append(RowError(row=i, field="active", message="must be TRUE/FALSE"))

            pr = get("priority")
            if pr and pr.upper() not in PRIORITIES:
                errors.append(RowError(row=i, field="priority", message="invalid priority (HIGH/MEDIUM/LOW)"))

            win = get("preferred_run_window")
            if win:
                try:
                    w = int(win)
                    if w not in RUN_WINDOWS:
                        errors.append(RowError(row=i, field="preferred_run_window", message="invalid run window (15–20)"))
                except Exception:
                    errors.append(RowError(row=i, field="preferred_run_window", message="invalid run window (15–20)"))

            # Normalized output row.
            out: dict[str, Any] = {k: get(k) for k in headers_norm}
            out["__rownum"] = i
            if out.get("gstin"):
                out["gstin"] = str(out["gstin"]).replace(" ", "").upper()
            if out.get("priority"):
                out["priority"] = str(out["priority"]).upper()
            if out.get("active"):
                out["active"] = str(out["active"]).upper()
            if out.get("client_email"):
                out["client_email"] = ", ".join(_split_emails(str(out["client_email"])))
            rows.append(out)

        logger.info(
            "client_master.parsed",
            sheetnames=wb.sheetnames,
            sheet=sheet,
            headers=headers_norm,
            max_row=int(ws.max_row or 0),
            processed_rows=int(processed),
            skipped_empty_rows=int(skipped_empty),
            errors=int(len(errors)),
        )
        return ParseResult(rows=rows, errors=errors)

    def parse_records(self) -> TypedParseResult:
        parsed = self.parse()
        if not parsed.ok:
            # Preserve existing RowError detail for deterministic troubleshooting.
            return TypedParseResult(records=[], errors=parsed.errors, headers=COLUMNS)

        typed: list[ClientRecord] = []
        type_errors: list[RowError] = []
        for r in parsed.rows:
            rownum = int(r.get("__rownum") or 0) or None
            try:
                preferred_run_window = int(str(r.get("preferred_run_window") or "18"))
            except Exception:
                preferred_run_window = 18
                type_errors.append(RowError(row=rownum or 0, field="preferred_run_window", message="invalid int"))
            active = str(r.get("active") or "FALSE").upper() == "TRUE"
            rec = ClientRecord(
                client_id=str(r.get("client_id") or "").strip() or None,
                client_name=str(r.get("client_name") or "").strip(),
                gstin=str(r.get("gstin") or "").strip().upper(),
                username=str(r.get("username") or "").strip(),
                password=str(r.get("password") or "") or None,
                client_email=str(r.get("client_email") or "").strip(),
                financial_year=str(r.get("financial_year") or "").strip(),
                active=active,
                priority=str(r.get("priority") or "").strip().upper(),
                tags=str(r.get("tags") or "").strip() or None,
                preferred_run_window=preferred_run_window,
                notes=str(r.get("notes") or "").strip() or None,
                rownum=rownum,
            )
            typed.append(rec)

        logger.info(
            "client_master.typed",
            headers=COLUMNS,
            records=int(len(typed)),
            type_errors=int(len(type_errors)),
        )
        return TypedParseResult(records=typed, errors=type_errors, headers=COLUMNS)

from __future__ import annotations

import time
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path
from typing import Awaitable, Callable

import redis.asyncio as redis
from aiogram import Router, types
from aiogram.filters import Command

from gst_automation.core.logging import get_logger
from gst_automation.core.settings import Settings
from gst_automation.clients.excel_parser import ClientMasterParser


logger = get_logger(__name__)

@dataclass(frozen=True, slots=True)
class CommandDeps:
    settings: Settings
    redis_client: redis.Redis


def _is_allowed(*, settings: Settings, message: types.Message) -> bool:
    user = message.from_user
    if not user:
        return False
    telegram_user_id = int(user.id)
    chat_id = int(message.chat.id)
    allowed = {int(x) for x in (settings.telegram_allowed_user_ids or [])}
    return (telegram_user_id in allowed) or (chat_id in allowed)


def _safe_reply(text: str) -> str:
    # Keep operator-facing errors short and non-leaky.
    return text.strip()[:3500]


Handler = Callable[[CommandDeps, types.Message], Awaitable[None]]


def _wrap_command(name: str, handler: Handler) -> Handler:
    async def _wrapped(deps: CommandDeps, message: types.Message) -> None:
        user = message.from_user
        telegram_user_id = int(user.id) if user else None
        chat_id = int(message.chat.id)
        start = time.monotonic()
        logger.info(
            "telegram.command.invoked",
            command=name,
            telegram_user_id=telegram_user_id,
            chat_id=chat_id,
            message_id=getattr(message, "message_id", None),
        )
        if not _is_allowed(settings=deps.settings, message=message):
            logger.warning("telegram.command.rejected", command=name, telegram_user_id=telegram_user_id, chat_id=chat_id)
            try:
                await message.answer("Not authorized.")
            except Exception:
                pass
            return
        try:
            await handler(deps, message)
            logger.info(
                "telegram.command.completed",
                command=name,
                telegram_user_id=telegram_user_id,
                chat_id=chat_id,
                duration_ms=int((time.monotonic() - start) * 1000),
            )
        except Exception as exc:  # noqa: BLE001
            logger.exception(
                "telegram.command.failed",
                command=name,
                exception_type=type(exc).__name__,
                exception_message=str(exc),
                telegram_user_id=telegram_user_id,
                chat_id=chat_id,
                duration_ms=int((time.monotonic() - start) * 1000),
                exc_info=True,
            )
            try:
                await message.answer(_safe_reply("Command failed. Check server logs for details."))
            except Exception:
                pass

    return _wrapped


async def _cmd_help(_deps: CommandDeps, message: types.Message) -> None:
    await message.answer(
        _safe_reply(
            "\n".join(
                [
                    "GST Automation Bot Commands:",
                    "",
                    "/start - activate bot",
                    "/ping - health check (pong)",
                    "/help - show this help",
                    "/status - bot + runtime status",
                    "/health - runtime health summary",
                    "/version - app version",
                    "/clients - summarize client_master.xlsx",
                    "/tasks - show running tasks (placeholder)",
                    "/download_gstr2b - start GSTR-2B flow (placeholder)",
                    "/restart_worker - restart workers (placeholder)",
                    "/logs - show safe runtime summary (placeholder)",
                    "/cancel_task - cancel a task (placeholder)",
                    "/whoami - show your telegram ids",
                ]
            )
        )
    )


async def _cmd_start(_deps: CommandDeps, message: types.Message) -> None:
    await message.answer("GST Bot active. Use /help for commands.")


async def _cmd_ping(_deps: CommandDeps, message: types.Message) -> None:
    await message.answer("pong")


async def _cmd_status(_deps: CommandDeps, message: types.Message) -> None:
    now = datetime.now(UTC).isoformat()
    await message.answer(_safe_reply(f"status=ok\npolling=active\nutc={now}"))


async def _cmd_health(deps: CommandDeps, message: types.Message) -> None:
    # Best-effort: redis ping only (no secrets).
    redis_ok = False
    try:
        redis_ok = bool(await deps.redis_client.ping())
    except Exception:
        redis_ok = False
    await message.answer(_safe_reply(f"health=ok\nredis_ok={str(redis_ok).lower()}"))


async def _cmd_version(_deps: CommandDeps, message: types.Message) -> None:
    try:
        from importlib.metadata import version

        v = version("gst-automation-platform")
    except Exception:
        v = "unknown"
    await message.answer(_safe_reply(f"version={v}"))


async def _cmd_clients(deps: CommandDeps, message: types.Message) -> None:
    # Deterministic: parse client_master.xlsx if present, otherwise placeholder.
    cwd = Path.cwd()
    data_dir = Path(deps.settings.data_dir)
    candidates = [
        Path("/app/client_master.xlsx"),
        cwd / "client_master.xlsx",
        data_dir / "client_master.xlsx",
        (cwd / "data") / "client_master.xlsx",
    ]
    existence = {str(p): bool(p.exists()) for p in candidates}
    path = next((p for p in candidates if p.exists()), None)
    logger.info(
        "telegram.clients.resolve",
        cwd=str(cwd),
        data_dir=str(data_dir),
        candidates=[str(p) for p in candidates],
        exists=existence,
        selected=str(path) if path else None,
    )
    if path is None:
        await message.answer(
            _safe_reply(
                "client_master.xlsx not found inside container.\n"
                "Searched:\n"
                + "\n".join([f"- {p}" for p in candidates])
            )
        )
        return
    logger.info("telegram.clients.parse.start", excel_path=str(path), exists=True)
    try:
        # Debug-only metadata: sheet names help diagnose mismatched templates.
        try:
            from openpyxl import load_workbook

            wb = load_workbook(filename=str(path), read_only=True, data_only=True)
            try:
                logger.info("telegram.clients.parse.sheets", excel_path=str(path), sheet_names=list(wb.sheetnames))
            finally:
                wb.close()
        except Exception as exc:  # noqa: BLE001
            logger.warning("telegram.clients.parse.sheets_failed", excel_path=str(path), err=str(exc))

        parsed = ClientMasterParser(path=path).parse_records()
    except Exception as exc:  # noqa: BLE001
        logger.exception(
            "telegram.clients.parse.exception",
            excel_path=str(path),
            exception_type=type(exc).__name__,
            exception_message=str(exc),
            exc_info=True,
        )
        raise
    if not parsed.ok:
        logger.info(
            "telegram.clients.parse.failed",
            excel_path=str(path),
            errors=int(len(parsed.errors)),
            records=int(len(getattr(parsed, "records", []) or [])),
        )
        top = parsed.errors[:5]
        details = "\n".join([f"- row {e.row} {e.field}: {e.message}" for e in top]) if top else ""
        await message.answer(
            _safe_reply(
                f"client_master.xlsx parse failed: {len(parsed.errors)} errors\n"
                + (details if details else "")
            )
        )
        return
    records = parsed.records
    logger.info("telegram.clients.parse.ok", excel_path=str(path), records=int(len(records)))

    lines: list[str] = [f"clients_ok=true", f"clients_count={len(records)}", ""]
    for idx, r in enumerate(records[:10], start=1):
        lines.append(f"{idx}. {r.client_name}")
        lines.append(f"   GSTIN: {r.gstin}")
        lines.append("")
    if len(records) > 10:
        lines.append(f"... ({len(records) - 10} more)")
    await message.answer(_safe_reply("\n".join(lines).strip()))


async def _cmd_tasks(_deps: CommandDeps, message: types.Message) -> None:
    await message.answer("tasks: placeholder (not implemented yet).")


async def _cmd_download_gstr2b(_deps: CommandDeps, message: types.Message) -> None:
    await message.answer("GSTR2B download workflow initiated (placeholder).")


async def _cmd_restart_worker(_deps: CommandDeps, message: types.Message) -> None:
    await message.answer("restart_worker: placeholder (operator must restart via docker compose).")


async def _cmd_logs(_deps: CommandDeps, message: types.Message) -> None:
    await message.answer("logs: placeholder (use docker compose logs -f api).")


async def _cmd_cancel_task(_deps: CommandDeps, message: types.Message) -> None:
    await message.answer("cancel_task: placeholder (not implemented yet).")


async def _cmd_whoami(_deps: CommandDeps, message: types.Message) -> None:
    user = message.from_user
    if not user:
        return
    uid = int(user.id)
    chat_id = int(message.chat.id)
    await message.answer(_safe_reply(f"telegram_user_id={uid}\nchat_id={chat_id}"))


def build_command_router(*, settings: Settings, redis_client: redis.Redis) -> Router:
    deps = CommandDeps(settings=settings, redis_client=redis_client)

    router = Router()

    def _h(name: str, func: Handler) -> Callable[[types.Message], Awaitable[None]]:
        wrapped = _wrap_command(name, func)

        async def _handler(message: types.Message) -> None:
            await wrapped(deps, message)

        return _handler

    router.message.register(_h("help", _cmd_help), Command("help"))
    router.message.register(_h("start", _cmd_start), Command("start"))
    router.message.register(_h("ping", _cmd_ping), Command("ping"))
    router.message.register(_h("status", _cmd_status), Command("status"))
    router.message.register(_h("health", _cmd_health), Command("health"))
    router.message.register(_h("version", _cmd_version), Command("version"))
    router.message.register(_h("clients", _cmd_clients), Command("clients"))
    router.message.register(_h("tasks", _cmd_tasks), Command("tasks"))
    router.message.register(_h("download_gstr2b", _cmd_download_gstr2b), Command("download_gstr2b"))
    router.message.register(_h("restart_worker", _cmd_restart_worker), Command("restart_worker"))
    router.message.register(_h("logs", _cmd_logs), Command("logs"))
    router.message.register(_h("cancel_task", _cmd_cancel_task), Command("cancel_task"))
    router.message.register(_h("whoami", _cmd_whoami), Command("whoami"))

    logger.info(
        "telegram.commands.registered",
        commands=[
            "help",
            "start",
            "ping",
            "status",
            "health",
            "version",
            "clients",
            "tasks",
            "download_gstr2b",
            "restart_worker",
            "logs",
            "cancel_task",
            "whoami",
        ],
    )
    return router

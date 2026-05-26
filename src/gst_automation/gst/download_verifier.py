from __future__ import annotations

import zipfile
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from gst_automation.archive.hashing import sha256_file


@dataclass(frozen=True, slots=True)
class DownloadVerificationResult:
    ok: bool
    classification: str  # ok/retryable/corrupt/permanent
    reasons: list[str]
    sha256_hex: str | None
    byte_size: int
    sheet_names: list[str]
    sampled_rows_any_sheet: int


@dataclass(frozen=True, slots=True)
class Gstr2bDownloadVerifier:
    min_size_bytes: int = 10_000
    min_sampled_rows_any_sheet: int = 2

    def verify(self, path: Path) -> DownloadVerificationResult:
        reasons: list[str] = []
        if not path.exists() or not path.is_file():
            return DownloadVerificationResult(
                ok=False,
                classification="retryable",
                reasons=["file missing"],
                sha256_hex=None,
                byte_size=0,
                sheet_names=[],
                sampled_rows_any_sheet=0,
            )

        size = int(path.stat().st_size)
        if size < self.min_size_bytes:
            # Heuristic only: a tiny but valid workbook (e.g. tests/minimal templates) should still pass
            # if it opens cleanly and contains the expected minimum rows.
            reasons.append(f"file too small ({size} bytes)")

        if path.suffix.lower() != ".xlsx":
            reasons.append("not an .xlsx file")

        # Corruption / partial file detection: validate zip structure first.
        try:
            with zipfile.ZipFile(path) as zf:
                bad = zf.testzip()
                if bad is not None:
                    reasons.append(f"zip corruption at {bad}")
        except zipfile.BadZipFile:
            return DownloadVerificationResult(
                ok=False,
                classification="corrupt",
                reasons=["bad zip (corrupt/partial download)"],
                sha256_hex=None,
                byte_size=size,
                sheet_names=[],
                sampled_rows_any_sheet=0,
            )
        except Exception as exc:  # noqa: BLE001
            return DownloadVerificationResult(
                ok=False,
                classification="retryable",
                reasons=[f"zip read failed: {exc}"],
                sha256_hex=None,
                byte_size=size,
                sheet_names=[],
                sampled_rows_any_sheet=0,
            )

        # XLSX open and basic content sanity.
        sheet_names: list[str] = []
        sampled_rows = 0
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
            sheet_names = list(wb.sheetnames)
            if not sheet_names:
                reasons.append("workbook has no sheets")
            for name in sheet_names[:5]:
                ws = wb[name]
                count = 0
                for _ in ws.iter_rows(max_row=50):
                    count += 1
                sampled_rows = max(sampled_rows, count)
            if sampled_rows < self.min_sampled_rows_any_sheet:
                reasons.append("not enough rows in any sheet (possible error download)")
        except Exception as exc:  # noqa: BLE001
            return DownloadVerificationResult(
                ok=False,
                classification="corrupt",
                reasons=[f"xlsx open failed: {exc}"],
                sha256_hex=None,
                byte_size=size,
                sheet_names=[],
                sampled_rows_any_sheet=0,
            )

        # If the workbook opened successfully and has enough rows, ignore the size-only heuristic.
        if sampled_rows >= self.min_sampled_rows_any_sheet:
            reasons = [r for r in reasons if not r.startswith("file too small")]

        sha = sha256_file(path)

        ok = not reasons
        classification = "ok" if ok else "permanent"
        # If it looks like a partial/corrupt download, classify retryable; otherwise treat as permanent validation failure.
        if not ok and any("too small" in r or "zip" in r or "open failed" in r for r in reasons):
            classification = "retryable"

        return DownloadVerificationResult(
            ok=ok,
            classification=classification,
            reasons=reasons,
            sha256_hex=sha,
            byte_size=size,
            sheet_names=sheet_names,
            sampled_rows_any_sheet=sampled_rows,
        )


def quarantine_file(*, src: Path, quarantine_dir: Path) -> Path:
    quarantine_dir.mkdir(parents=True, exist_ok=True)
    target = quarantine_dir / src.name
    if target.exists():
        target = quarantine_dir / f"{src.stem}_{src.stat().st_size}_{src.suffix.lstrip('.')}"
        target = target.with_suffix(src.suffix)
    src.replace(target)
    return target

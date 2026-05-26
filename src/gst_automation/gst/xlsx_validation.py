from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook


@dataclass(frozen=True, slots=True)
class XlsxValidationResult:
    ok: bool
    sheet_names: list[str]
    min_rows_ok: bool
    byte_size: int


@dataclass(frozen=True, slots=True)
class XlsxValidator:
    min_size_bytes: int = 10_000
    min_rows_any_sheet: int = 1

    def validate(self, path: Path) -> XlsxValidationResult:
        size = path.stat().st_size
        if size < self.min_size_bytes:
            return XlsxValidationResult(ok=False, sheet_names=[], min_rows_ok=False, byte_size=size)
        wb = load_workbook(path, read_only=True, data_only=True)
        sheets = list(wb.sheetnames)
        min_rows_ok = False
        for name in sheets[:5]:
            ws = wb[name]
            # openpyxl read_only row count is approximate; sample first N rows.
            count = 0
            for _ in ws.iter_rows(max_row=50):
                count += 1
            if count >= self.min_rows_any_sheet:
                min_rows_ok = True
                break
        ok = bool(sheets) and min_rows_ok
        return XlsxValidationResult(ok=ok, sheet_names=sheets, min_rows_ok=min_rows_ok, byte_size=size)

